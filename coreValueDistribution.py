import openpyxl
import collections
files = "NDCC TaMS NDCS TaAU ThAU ThMS CoMH CoGe".split()

wbE = openpyxl.Workbook()
for file in files:
    wsE = wbE.create_sheet(file)
    for x in range(2,11):
    # for x in range(1):
        filename = f"./result/kxCoreValue/{file}-{x}-coreValue.txt"
        # filename = f"./result/kCoreValue/{file}-coreValue.txt"
        data=[]
        with open(filename, "r") as f:
            for line in f:
                data.append(line)
        cE = data[0].split()
        if len(cE) == 0:
            continue
        cE = [int(info) for info in cE]
        
        # Calculate the maximum element in cE and cV
        cE_max = max(cE)
        cE_dis = [0 for i in range(cE_max+1)]
        for info in cE:
            cE_dis[info] += 1
        for i in range(2, cE_max+1):
            cE_dis[i]+=cE_dis[i-1]
        # cE_pro=[]
        # for i in range(1,cE_max+1):
        #     cE_pro.append(cE_dis[i]/cE_dis[cE_max])
        cE_dis.pop(0)
        for i in range(cE_max):
            wsE.cell(row=i+1, column=x-1).value = cE_dis[i]
            # wsE.cell(row=i+1, column=1).value = cE_dis[i]
        print(f"{file} done")
wbE.remove(wbE["Sheet"])
wbE.save("./result/kxCoreValue.xlsx")
